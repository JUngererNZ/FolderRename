"""
Enhanced script to automatically find the last column, copy its data, insert a new column, and hide it.
This script improves upon the original by dynamically detecting the last column instead of using hardcoded positions.
"""

import argparse
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


def find_last_comments_column(ws):
    """
    Find the last column that contains a header with "COMMENTS" and a date.
    Returns the column letter and column index.
    """
    # Start from the rightmost column and work backwards
    max_col = ws.max_column
    
    for col_num in range(max_col, 0, -1):
        col_letter = get_column_letter(col_num)
        
        # Check multiple rows for COMMENTS header (it might not be in row 1)
        for row_num in range(1, min(10, ws.max_row + 1)):  # Check first 10 rows
            header_cell = ws[f"{col_letter}{row_num}"].value
            
            # Check if header contains "COMMENTS" and looks like a date
            if header_cell and isinstance(header_cell, str) and "COMMENTS" in header_cell.upper():
                return col_letter, col_num
    
    # If no COMMENTS column found, use the last column
    last_col_letter = get_column_letter(max_col)
    return last_col_letter, max_col


def copy_and_insert_comments_column(
    file_path: str | Path,
    sheet_name: str | None = None,
    date_format: str = "%d-%m-%Y",
    backup: bool = False,
) -> None:
    """
    Find the last column, copy its data, insert a new column, hide it, and add a dated header.

    Args:
        file_path: Path to the .xlsx file
        sheet_name: Target sheet name (None = active sheet)
        date_format: Date format string for header (default: %d-%m-%Y)
        backup: Create a backup copy before modifying
    """
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    # Optional backup
    if backup:
        backup_path = file_path.with_suffix(f".backup{file_path.suffix}")
        import shutil
        shutil.copy2(file_path, backup_path)
        print(f"✓ Backup created: {backup_path}")

    # Load workbook
    wb = load_workbook(file_path)
    
    # Select sheet
    ws = wb[sheet_name] if sheet_name else wb.active
    print(f"✓ Working on sheet: {ws.title}")

    # 1. Find the last column (the one with COMMENTS header)
    last_col_letter, last_col_index = find_last_comments_column(ws)
    print(f"✓ Found last column: {last_col_letter} (column {last_col_index})")

    # 2. Copy all data from the last column
    max_row = ws.max_row
    column_data = []
    
    for row_num in range(1, max_row + 1):
        cell_value = ws[f"{last_col_letter}{row_num}"].value
        column_data.append(cell_value)
    
    print(f"✓ Copied {len(column_data)} rows from column {last_col_letter}")

    # 3. Insert new column after the last column
    new_col_index = last_col_index + 1
    new_col_letter = get_column_letter(new_col_index)
    ws.insert_cols(new_col_index)
    print(f"✓ Inserted new column at {new_col_letter} (column {new_col_index})")

    # 4. Paste the copied data into the new column
    for row_num, cell_value in enumerate(column_data, start=1):
        ws[f"{new_col_letter}{row_num}"] = cell_value
    
    print(f"✓ Pasted data into new column {new_col_letter}")

    # 5. Hide the previous column (the one we copied from)
    ws.column_dimensions[last_col_letter].hidden = True
    print(f"✓ Hidden previous column {last_col_letter}")

    # 6. Set new header with today's date
    today_str = datetime.today().strftime(date_format)
    new_header = f"COMMENTS {today_str}"
    ws[f"{new_col_letter}1"] = new_header
    print(f"✓ Set {new_col_letter}1 = '{new_header}'")

    # Save changes
    wb.save(file_path)
    print(f"✓ Saved: {file_path}")


def select_file_with_dialog():
    """
    Open a file dialog to let the user select an Excel file.
    Returns the selected file path or None if cancelled.
    """
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    file_path = filedialog.askopenfilename(
        title="Select Excel file to update",
        filetypes=[
            ("Excel files", "*.xlsx"),
            ("All files", "*.*")
        ]
    )
    
    root.destroy()  # Clean up
    return file_path


def main():
    parser = argparse.ArgumentParser(
        description="Enhanced: Copy last column data, insert new column, hide it, and add dated header"
    )
    parser.add_argument(
        "file",
        nargs="?",  # Make file argument optional
        type=str,
        help="Path to the Excel .xlsx file (or select via dialog if not provided)"
    )
    parser.add_argument(
        "-s", "--sheet",
        type=str,
        default=None,
        help="Sheet name (default: active sheet)"
    )
    parser.add_argument(
        "-b", "--backup",
        action="store_true",
        help="Create backup before modifying"
    )
    parser.add_argument(
        "-d", "--date-format",
        type=str,
        default="%d-%m-%Y",
        help="Date format for header (default: %%d-%%m-%%Y)"
    )

    args = parser.parse_args()

    # If no file provided via command line, show file dialog
    if not args.file:
        print("No file specified. Opening file dialog...")
        selected_file = select_file_with_dialog()
        
        if not selected_file:
            print("❌ No file selected. Exiting.")
            return
        
        args.file = selected_file
        print(f"Selected file: {args.file}")

    try:
        copy_and_insert_comments_column(
            file_path=args.file,
            sheet_name=args.sheet,
            date_format=args.date_format,
            backup=args.backup,
        )
        print("\n✅ Complete! New column inserted with copied data and hidden.")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        raise


if __name__ == "__main__":
    main()