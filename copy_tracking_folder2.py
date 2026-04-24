#! C:\Users\Jason\AppData\Local\Programs\Python\Python314\python.exe
import os
import shutil
import re
import sys
import json
from datetime import datetime
from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =========================
# Shared Configuration
# =========================
BASE_DIR = r"C:\Users\Jason\FML Freight Solutions\FML Doc Share - Documents\TRACKING\APRIL 2026"
TODAY_STR = datetime.today().strftime("%d-%m-%Y")
HEADER_PATTERN = re.compile(r"^COMMENTS \d{2}-\d{2}-\d{4}$")

PREFIXES = {
    "BARTRAC - CONGO TRACKING":                   f"BARTRAC - CONGO TRACKING {TODAY_STR}.xlsx",
    "BARTRAC - KAMOA TRACKING AS OF":             f"BARTRAC - KAMOA TRACKING AS OF {TODAY_STR}.xlsx",
    "BARTRAC - KCC TRACKING AS OF":               f"BARTRAC - KCC TRACKING AS OF {TODAY_STR}.xlsx",
    "BARTRAC - TRACKING - FML BONDED FACILITY -": f"BARTRAC - TRACKING - FML BONDED FACILITY - {TODAY_STR}.xlsx",
    "FML-KANU - ALLAN - TRACKING AS OF":          f"FML-KANU - ALLAN - TRACKING AS OF {TODAY_STR}.xlsx",
    "BARTRAC - ERG TRACKING":                     f"BARTRAC - ERG TRACKING {TODAY_STR}.xlsx",
    "BARTRAC - SURYA MINES":                      f"BARTRAC - SURYA MINES {TODAY_STR}.xlsx",
}

def load_config(config_file="tracking_workflow_config.json"):
    with open(config_file, "r") as f:
        return json.load(f)

def duplicate_comments_column(file_path: Path, sheet_name: str, header_row: int) -> bool:
    try:
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            print(f"  [ERROR] Sheet '{sheet_name}' not found in {file_path.name}")
            return False

        ws = wb[sheet_name]
        target_col = None

        for col_idx in range(1, ws.max_column + 1):
            header_value = ws.cell(row=header_row, column=col_idx).value
            if isinstance(header_value, str) and HEADER_PATTERN.match(header_value.strip()):
                target_col = col_idx

        if target_col is None:
            print(f"  [ERROR] No matching 'COMMENTS' header found in row {header_row}.")
            return False

        column_data = []
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=target_col)
            column_data.append({
                "row": row_idx, "value": cell.value, "font": copy(cell.font),
                "fill": copy(cell.fill), "border": copy(cell.border), "alignment": copy(cell.alignment),
                "number_format": cell.number_format, "protection": copy(cell.protection),
                "comment": copy(cell.comment) if cell.comment else None,
                "hyperlink": copy(cell.hyperlink) if cell.hyperlink else None, "has_style": cell.has_style,
            })

        target_letter = get_column_letter(target_col)
        original_width = ws.column_dimensions[target_letter].width

        merged_ranges_to_duplicate = []
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_col <= target_col <= merged_range.max_col:
                width = merged_range.max_col - merged_range.min_col
                offset = target_col - merged_range.min_col
                merged_ranges_to_duplicate.append((merged_range.min_row, target_col + 1 - offset, merged_range.max_row, (target_col + 1 - offset) + width))

        ws.insert_cols(target_col + 1, 1)
        for item in column_data:
            new_cell = ws.cell(row=item["row"], column=target_col + 1)
            new_cell.value = item["value"]
            if item["has_style"]:
                new_cell.font, new_cell.fill, new_cell.border, new_cell.alignment, new_cell.number_format, new_cell.protection = \
                copy(item["font"]), copy(item["fill"]), copy(item["border"]), copy(item["alignment"]), item["number_format"], copy(item["protection"])
            if item["comment"]: new_cell.comment = item["comment"]
            if item["hyperlink"]: new_cell._hyperlink = item["hyperlink"]

        ws.column_dimensions[get_column_letter(target_col + 1)].width = original_width
        for min_row, min_col, max_row, max_col in merged_ranges_to_duplicate:
            try: ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
            except ValueError: pass

        ws.column_dimensions[get_column_letter(target_col)].hidden = True
        wb.save(file_path)
        print(f"  [SUCCESS] Column duplicated for {file_path.name}")
        return True
    except Exception as e:
        print(f"  [ERROR] Failure updating {file_path.name}: {e}")
        return False

def main():
    config = load_config()
    excel_ops = config["workflow"]["configurations"]["excel_operations"]["targets"]

    print("--- STEP 1: Copying Folders ---")
    folders = [ (datetime.strptime(d, "%d-%m-%Y"), os.path.join(BASE_DIR, d)) 
                for d in os.listdir(BASE_DIR) if os.path.isdir(os.path.join(BASE_DIR, d)) ]
    
    if not folders: sys.exit("[FATAL] No folders found.")
    latest_folder = sorted(folders)[-1][1]
    target_path = os.path.join(BASE_DIR, TODAY_STR)
    
    if not os.path.exists(target_path): shutil.copytree(latest_folder, target_path)

    print("\n--- STEP 2: Renaming Files ---")
    for fname in os.listdir(target_path):
        for prefix, new_name in PREFIXES.items():
            if fname.startswith(prefix):
                os.rename(os.path.join(target_path, fname), os.path.join(target_path, new_name))
                break

    print("\n--- STEP 3: Excel Operations ---")
    found_files = {op_key: Path(target_path) / fname for fname in os.listdir(target_path) 
                   for op_key, op_data in excel_ops.items() if op_data["search_key"].upper() in fname.upper()}

    for op_key, op_data in excel_ops.items():
        print(f"Updating {op_key}...")
        if found_files.get(op_key):
            duplicate_comments_column(found_files[op_key], op_data["sheet_name"], op_data["header_row"])

if __name__ == "__main__":
    main()