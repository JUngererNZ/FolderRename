"""
Insert-and-hide-a-column.py

Python/openpyxl replication of Script-2.osts (the Office Script recorded in Excel).

Inserts a new blank column at a specified position, hides it, then writes a dated
'COMMENTS DD-MM-YYYY' header into the next column — matching the manual workflow
performed weekly on BARTRAC tracking workbooks.

Usage:
    python Insert-and-hide-a-column.py <file.xlsx> [options]

    Options:
        -s, --sheet           Target sheet name (default: active sheet)
        -i, --insert-column   Column to insert at (default: MB)
        -c, --header-column   Column to write dated header into (default: MC)
        -d, --date-format     Date format string (default: %d-%m-%Y)
        -b, --backup          Create a .backup.xlsx copy before modifying

Examples:
    python Insert-and-hide-a-column.py "BARTRAC - KCC TRACKING.xlsx" --backup
    python Insert-and-hide-a-column.py "BARTRAC - KCC TRACKING.xlsx" -s "ENROUTE SITE" -i MB -c MC

Difference from duplicate_hide_mk_column.py:
    - This script inserts a blank hidden column and adds a new dated header
    - The other script duplicates an existing COMMENTS column then hides the copy
    - These are complementary steps: run duplicate_hide_mk_column.py first to
      preserve last week's comments, then this script to create the new week's column

Limitations:
    - openpyxl does not preserve macros or some advanced Excel formatting
    - Column positions (MB/MC) must match the current state of the workbook
    - Does not read tracking_workflow_config.json — sheet and column args must be
      supplied manually per client
"""

"""
Insert and hide a column, then add a dated comment header.
Replicates the behavior of Script-2.osts for local Excel files.
"""

import argparse
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def insert_comments_column(
    file_path: str | Path,
    sheet_name: str | None = None,
    insert_column: str = "MB",
    header_column: str = "MC",
    date_format: str = "%d-%m-%Y",
    backup: bool = False,
) -> None:
    """
    Insert a new column, hide it, and add a dated header.

    Args:
        file_path: Path to the .xlsx file
        sheet_name: Target sheet name (None = active sheet)
        insert_column: Column letter to insert at (default: MB)
        header_column: Column letter for the header cell (default: MC)
        date_format: Date format string for header (default: %d-%m-%Y)
        backup: Create a backup copy before modifying
    """
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    # Optional backup
    if backup:
        backup_path = file_path.with_suffix(f".backup{file_path.suffix}")
        import shutil
        shutil.copy2(file_path, backup_path)
        print(f"✓ Backup created: {backup_path}")

    # Load workbook
    wb = load_workbook(file_path)
    
    # Select sheet
    ws = wb[sheet_name] if sheet_name else wb.active
    print(f"✓ Working on sheet: {ws.title}")

    # 1. Insert column at specified position (shifts right)
    col_index = column_index_from_string(insert_column)
    ws.insert_cols(col_index)
    print(f"✓ Inserted column at {insert_column}")

    # 2. Hide the inserted column
    ws.column_dimensions[insert_column].hidden = True
    print(f"✓ Hidden column {insert_column}")

    # 3. Set header with today's date
    today_str = datetime.today().strftime(date_format)
    header_cell = f"{header_column}1"
    ws[header_cell] = f"COMMENTS {today_str}"
    print(f"✓ Set {header_cell} = 'COMMENTS {today_str}'")

    # Save changes
    wb.save(file_path)
    print(f"✓ Saved: {file_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Insert and hide a column with dated comment header"
    )
    parser.add_argument(
        "file",
        type=str,
        help="Path to the Excel .xlsx file"
    )
    parser.add_argument(
        "-s", "--sheet",
        type=str,
        default=None,
        help="Sheet name (default: active sheet)"
    )
    parser.add_argument(
        "-i", "--insert-column",
        type=str,
        default="MB",
        help="Column to insert at (default: MB)"
    )
    parser.add_argument(
        "-c", "--header-column",
        type=str,
        default="MC",
        help="Column for header cell (default: MC)"
    )
    parser.add_argument(
        "-b", "--backup",
        action="store_true",
        help="Create backup before modifying"
    )
    parser.add_argument(
        "-d", "--date-format",
        type=str,
        default="%d-%m-%Y",
        help="Date format for header (default: %%d-%%m-%%Y)"
    )

    args = parser.parse_args()

    try:
        insert_comments_column(
            file_path=args.file,
            sheet_name=args.sheet,
            insert_column=args.insert_column,
            header_column=args.header_column,
            date_format=args.date_format,
            backup=args.backup,
        )
        print("\n✅ Complete!")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        raise


if __name__ == "__main__":
    main()
