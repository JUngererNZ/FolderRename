import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
import re

def find_last_comments_column(ws):
    """
    Find the last column in row 1 that matches the pattern 'COMMENTS DD-MM-YYYY'.
    Returns the column letter if found, else None.
    """
    pattern = re.compile(r'COMMENTS \d{2}-\d{2}-\d{4}')
    last_col = None
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value and pattern.match(str(cell_value).strip()):
            last_col = get_column_letter(col)
    return last_col

def duplicate_and_hide_column(file_path):
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active  # First/active sheet
        
        # Find the last column with COMMENTS header
        target_col = find_last_comments_column(ws)
        if not target_col:
            print("Error: No column with 'COMMENTS DD-MM-YYYY' header found in row 1.")
            return
        
        print(f"Found target column: {target_col}")
        
        # Get column index
        target_index = column_index_from_string(target_col)
        
        # Insert new column before the target column (at target_index position)
        ws.insert_cols(target_index)
        new_col_letter = get_column_letter(target_index)
        
        print(f"Inserted new column: {new_col_letter}")
        
        # Copy data from the shifted original column (now at target_index + 1) to the new column (at target_index)
        max_row = ws.max_row
        for row in range(1, max_row + 1):
            original_value = ws.cell(row=row, column=target_index + 1).value
            ws.cell(row=row, column=target_index).value = original_value
        
        print(f"Copied data from shifted {get_column_letter(target_index + 1)} to {new_col_letter}")
        
        # Hide the new column
        ws.column_dimensions[new_col_letter].hidden = True
        
        print(f"Hidden column: {new_col_letter}")
        
        # Save the workbook
        wb.save(file_path)
        print("Workbook saved successfully.")
        
    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    file_path = r"C:\Users\Jason\OneDrive - FML Freight Solutions\FML-PROJECTS\FolderRename\BARTRAC - KCC TRACKING AS OF 19-03-2026.xlsx"
    duplicate_and_hide_column(file_path)